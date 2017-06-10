from django.shortcuts import render,redirect,get_object_or_404, render_to_response
#encoding:utf-8
# Create your views here.
from personas.models import Persona
from django.core.urlresolvers import reverse_lazy, reverse
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.list import ListView
from django.views.generic.base import TemplateView, View
from django.views.generic.detail import DetailView


from django.conf import settings
from .forms import PersonaCreateForm, CttoUpdateForm, EdpUpdateForm, EdpCreateForm, CttaUpdateForm, OdcUpdateForm, OdcCreateForm,\
ItemOdcFormSet, ItemCttoFormSet, AportesCttoFormSet, MultasPerClaveCttoFormSet, PersonalProyUpdateForm, PersonalCttaUpdateForm

#Workbook nos permite crear libros en excel

from openpyxl import Workbook, load_workbook

#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse

from django.db.models import Avg, Max, Min, Count, Sum
from django.db import transaction

from django.core.exceptions import ObjectDoesNotExist # para cuando exista moneda

from datetime import timedelta, datetime, date, timedelta

import date_converter

#-------------------------------------------------------------------------

#from django.shortcuts import render_to_response
from django.shortcuts import render
from django.http import HttpResponseBadRequest, JsonResponse
from django import forms
from django.template import RequestContext
import django_excel as excel
from .models import Question, Choice, Area, Ceco, Mdte, Ctta, Ctto, Edp, Odc, Monedas, ItemOdc, ItemCtto,AportesCtto,PersonalProyecto,PersonalCtta

# No longer you need the following import statements if you use pyexcel >=0.2.2
import pyexcel.ext.xls
import pyexcel.ext.xlsx
import pyexcel.ext.ods3   # noqa

import unittest
from numerosletras import number_to_letter
import pdb
import locale

locale.setlocale(locale.LC_ALL,"")

from docxtpl import DocxTemplate, RichText
import time











class UploadFileForm(forms.Form):
    file = forms.FileField()

# Create your views here.
def upload(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            filehandle = request.FILES['file']
            return excel.make_response(filehandle.get_sheet(), "csv",
                                       file_name="download")
    else:
        form = UploadFileForm()
    return render_to_response(
        'upload_form.html',
        {
            'form': form,
            'title': 'Excel file upload and download example',
            'header': ('Please choose any excel file ' +
                       'from your cloned repository:')
        },
        context_instance=RequestContext(request))

def import_sheet(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
        if form.is_valid():
            request.FILES['file'].save_to_database(
                name_columns_by_row=0,
                model=Ceco,
                mapdict=['IdCeco', 'IdAreas', 'CodCeco', 'NomCeco', 'Budget'])
            return HttpResponse("OK")
        else:
            return HttpResponseBadRequest()
    else:
        form = UploadFileForm()
    return render_to_response('upload_form.html',
                              {'form': form},
                              context_instance=RequestContext(request))


def import_data(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        def Ceco_func(row):
            q = Area.objects.filter(IdAreas=row[1])[0]
            row[1] = q
            return row
        def Ctto_func(row):
            print ('en Ctto')
            print (row)
            q1 = Ctta.objects.filter(IdCtta=row[5])[0]
            row[5] = q1
            q2 = Ceco.objects.filter(IdCeco=row[9])[0]
            row[9] = q2
            q3 = Mdte.objects.filter(IdMandante=row[11])[0]
            row[11] = q3
            return row
        def Edp_func(row):
            print ('en EDP')
            print (row)
            q4 = Ctto.objects.filter(IdCtto=row[1])[0]
            row[1] = q4
            return row
        def Odc_func(row):
            print(row)
            q5 = Ceco.objects.filter(IdCeco=row[2])[0]
            row[2] = q5
            q6 = Ctto.objects.filter(IdCtto=row[3])[0]
            row[3] = q6
            print('q5 y q6')
            print(q5,q6)
            return row

        if form.is_valid():
            request.FILES['file'].save_book_to_database(
                models=[Area, Ceco, Mdte, Ctta, Ctto, Edp, Odc],
                initializers=[None, Ceco_func, None, None,Ctto_func,Edp_func, Odc_func],
                mapdicts=[
                    ['IdAreas', 'NomArea', 'CodArea'],
                    ['IdCeco', 'IdAreas', 'CodCeco', 'NomCeco', 'Budget'],
                    ['IdMandante', 'NomMandte', 'DirecMandte', 'RutMandte'],
                    ['IdCtta', 'NomCtta', 'DirCtta', 'RutCtta'],
                    ['IdCtto', 'NumCtto', 'DescCtto', 'MonedaCtto', 'ValorCtto', 'IdCtta', 'EstCtto', 'FechIniCtto', 'FechTerCtto', 'IdCecoCtto', 'CordCtto', 'IdMandante',\
                    'Carpeta','TipoServ', 'AjusteCom','AjustNumEDP','AjustValEDP','AdjudicCtto','LocalCtto','TerrenCtto','SeguroCtto'],
                    ['IdEDP', 'IdCtto', 'NumEDP', 'ValEDP', 'PeriodEDP', 'DevAntEDP', 'RetEDP', 'DevRet', 'Estado', 'FactEDP'],
                    ['IdODC', 'NumODC', 'IdCecoODC', 'IdCtto', 'FechT_ODC', 'ValorODC', 'DescripODC']]

            )
            return HttpResponse("OK", status=200)
        else:
            return HttpResponseBadRequest()
    else:
        form = UploadFileForm()
    return render(request,
        'upload_form.html',
        {
            'form': form,
            'title': 'Import excel data into database example',
            'header': 'Please upload sample-data.xls:'
        })


def import_EDP_ODC(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        def Edp_func(row):
            print ('en EDP')
            print (row)
            q4 = Ctto.objects.filter(IdCtto=row[1])[0]
            row[1] = q4
            return row
        def Odc_func(row):
            print(row)
            q5 = Ceco.objects.filter(IdCeco=row[2])[0]
            row[2] = q5
            q6 = Ctto.objects.filter(IdCtto=row[3])[0]
            row[3] = q6
            print('q5 y q6')
            print(q5,q6)
            return row

        if form.is_valid():
            print('valido')
            request.FILES['file'].save_book_to_database(
                models=[Edp, Odc],
                initializers=[Edp_func, Odc_func],
                mapdicts=[
                    ['IdEDP', 'IdCtto', 'NumEDP', 'ValEDP', 'PeriodEDP', 'DevAntEDP', 'RetEDP', 'DevRet', 'Estado', 'FactEDP'],
                    ['IdODC', 'NumODC', 'IdCecoODC', 'IdCtto', 'FechT_ODC', 'ValorODC', 'DescripODC']]
            )
            return HttpResponse("OK", status=200)
        else:
            return HttpResponseBadRequest()
    else:
        form = UploadFileForm()
    return render(request,
        'upload_form.html',
        {
            'form': form,
            'title': 'Import excel data into database example',
            'header': 'Please upload sample-data.xls:'
        })
















def export_r5(request):
            column_names = ['IdCtto','NumCtto', 'DescCtto']
            exp_ctto = Ctto.objects.filter()



            e = Ctto.objects.filter()
            print(e)

            #q5 = Mdte.objects.filter(exp_ctto)
            #q6 = Ctto.objects.select_related().get(id=2)


            #print(q6.NomMandte)

            #exp_ctto = Ctto.objects.all()


            return excel.make_response_from_query_sets(
                exp_ctto,
                column_names,
                'xls',
                file_name="custom"
            )






def export_data(request, atype):
    if atype == "sheet":
        return excel.make_response_from_a_table(
            Question, 'xls', file_name="sheet")
    elif atype == "book":
        return excel.make_response_from_tables(
            [Question, Choice], 'xls', file_name="book")
    elif atype == "custom":
        question = Question.objects.get(slug='ide')
        query_sets = Choice.objects.filter(question=question)
        column_names = ['choice_text', 'id', 'votes']
        return excel.make_response_from_query_sets(
            query_sets,
            column_names,
            'xls',
            file_name="custom"
        )
    else:
        return HttpResponseBadRequest(
            "Bad request. please put one of these " +
            "in your url suffix: sheet, book or custom")










#-------------------------------------------------------------------------------









def prueba(request):
    CTTOS = Ctto.objects.all()
    ODC = Odc.objects.all()
    EDP = Edp.objects.all()



    #Aux = Ctto.objects.values_list('id','IdCtto','NumCtto')
    #Aux = Ctto.objects.get(id=3)
    #Aux2 = Aux.IdCtta.NomCtta

    #Aux = list(Odc.objects.filter(IdCtto__id='3').values_list('ValorODC',flat=True))
    Aux = Odc.objects.filter(IdCtto__id=3).aggregate(Sum('ValorODC'))

    #Aux2= Ctto.objects.filter(id=3)
    #Aux = sumarLista(Aux)
    #Aux2 = Aux[0]
    Aux2 =Aux['ValorODC__sum']

    print('Aux', Aux)
    print('Aux2', Aux2)
    html="<html><body> el valor es Aux : %s  Aux2 : </body></html>" % Aux, Aux2

    return HttpResponse(html)








#Nuestra clase hereda de la vista generica TemplateView

def fac(moneda):
    DolarProyecto=680
    valor = 0
    try:
        valor = Monedas.objects.get(NomMoneda=moneda).ValorMoneda
        valor = valor/DolarProyecto

    except ObjectDoesNotExist:
        valor =0

    return valor

def fechaPalabra(fecha):

    formato_fecha = "%Y-%m-%d"
    mes = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
    dia_semana = {0:"lunes",1:"martes",2:"miércoles",3:"jueves",4:"viernes",5:"sábado",6:"domingo"}

    valor = ""
    try:
        valor = datetime.strptime(str(fecha),formato_fecha)
        valor =  str(valor.day) + " de " + str(mes[valor.month]) + " de " + str(valor.year)
    except:
        valor = ""
    return valor


def Plazodiaz(Finicio,Ftermino):

    formato_fecha = "%Y-%m-%d"
    valor = 0

    try:
        fecha_inicial = datetime.strptime(str(Finicio),formato_fecha)
        fecha_final = datetime.strptime(str(Ftermino),formato_fecha)
        valordias = fecha_final - fecha_inicial
        valor = valordias.days
    except:
        valor = 0

    return valor













class ReportePersonasExcel(TemplateView):

    #Usamos el metodo get para generar el archivo excel
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        CTTOS = Ctto.objects.all()
        ODC = Odc.objects.all()
        EDP = Edp.objects.all()
        Estatustxt = [ 'Solicitados en la semana','No iniciados','En preparación','En licitación','En evaluación','En negociación / Adjudicación','Adjudicado','En emisión','En firmas','En ejecución','Servicio terminado','En cierre','Cerrado','Servicio suspendido','Solicitud anulada','Solicitud diferida o postergada']



        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        #ws['B1'] = 'REPORTE DE PERSONAS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        #ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A1'] = 'Centro Costo'
        ws['B1'] = 'Cuenta'
        ws['C1'] = 'Mandante'
        ws['D1'] = 'Tipo'
        ws['E1'] = 'N° Ctto.'
        ws['F1'] = 'Descripcion Servicio'
        ws['G1'] = 'Contratista'
        ws['H1'] = 'Rut Contratista'
        ws['I1'] = 'Fecha Ini Ctto'
        ws['J1'] = 'Fecha Term Ctto'
        ws['K1'] = 'Estatus'
        ws['L1'] = 'Area'
        ws['M1'] = 'Cuenta'
        ws['N1'] = 'Descrip-Cuenta'
        ws['O1'] = 'Moneda Ctto'
        ws['P1'] = 'Valor Inicial'
        ws['Q1'] = 'Ajuste Commit Proyecto'
        ws['R1'] = 'EDP Ini Proy'
        ws['S1'] = 'EDP Ajust Proy'
        ws['T1'] = 'Adjudicación'
        ws['U1'] = 'Local'
        ws['V1'] = 'Terreno'
        ws['W1'] = 'Seguro'
        ws['X1'] = 'Valor ODC'
        ws['Y1'] = 'Valor EDP'
        ws['Z1'] = 'Val Actual Ctto'
        ws['AA1'] = 'Commitment Aprobado'
        ws['AB1'] = 'EDP Pagados Proy'
        ws['AC1'] = 'EDP Pagados Proy (USD)'
        ws['AD1'] = 'Commitment (USD)'
        ws['AE1'] = 'Commitment To Go (USD)'
        ws['AF1'] = 'Termino Actualizado'
        ws['AG1'] = 'Fecha Sol Ultima ODC'
        ws['AH1'] = 'Fecha Aprob Ultimo ODC'
        ws['AI1'] = 'Fecha Present Ultimo EDP'
        ws['AJ1'] = 'Fecha Aprob Ultimo EDP'
        ws['AK1'] = 'Fecha Periodo Ultimo EDP'
        ws['AL1'] = 'Fecha Solicitud Ctto'
        ws['AM1'] = 'Fecha Aprob Ctto'
        ws['AN1'] = 'Rut Ctta'
        ws['AO1'] = 'Observ Cttos'
        ws['AP1'] = 'Giro Ctta'
        ws['AQ1'] = 'Direccion Ctta'
        ws['AR1'] = 'Comuna Ctta'
        ws['AS1'] = 'Ciudad Ctta'
        ws['AT1'] = 'Tipo Prov'
        ws['AU1'] = 'Valor Inicio (USD)'
        ws['AV1'] = 'Valor ODC (USD)'
        ws['AW1'] = 'Num Ctto Descrip'
        ws['AX1'] = 'Estatus Texto'
        ws['AY1'] = 'Coordinador Tec NU'
        ws['AZ1'] = 'Administrador Ctta'
        ws['BA1'] = 'Cargo Administrador Ctta'
        ws['BB1'] = 'Correo Admnistrador Ctta'
        ws['BC1'] = 'Telefono Administrdaor Ctta'
        ws['BD1'] = 'Suma Item Ctto'
        ws['BE1'] = 'Suma Item Odc'
        ws['BF1'] = 'Suma Item CTT+ODC-Adj'
        ws['BG1'] = 'Suma Item Ctto (USD)'
        ws['BH1'] = 'Suma Item Odc (USD)'
        ws['BI1'] = 'Suma Item Total (USD)'
        ws['BJ1'] = 'Suma Retención'
        ws['BK1'] = 'Suma Dev Retención'
        ws['BL1'] = 'Saldo Retención (USD)'
        ws['BM1'] = 'factor (Moneda)'

        cont=2
        valcttoAct = 0
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for ctto in CTTOS:
            ws.cell(row=cont,column=1).value = ctto.IdCecoCtto.IdAreas.CodArea
            ws.cell(row=cont,column=2).value = ctto.IdCecoCtto.CodCeco
            ws.cell(row=cont,column=3).value = ctto.IdMandante.NomMandte
            ws.cell(row=cont,column=4).value = ctto.TipoServ
            ws.cell(row=cont,column=5).value = ctto.NumCtto
            ws.cell(row=cont,column=6).value = ctto.DescCtto
            ws.cell(row=cont,column=7).value = ctto.IdCtta.NomCtta
            ws.cell(row=cont,column=8).value = ctto.IdCtta.RutCtta
            ws.cell(row=cont,column=9).value = ctto.FechIniCtto
            ws.cell(row=cont,column=10).value = ctto.FechTerCtto
            ws.cell(row=cont,column=11).value = ctto.EstCtto
            ws.cell(row=cont,column=12).value = ctto.IdCecoCtto.IdAreas.NomArea
            ws.cell(row=cont,column=13).value = ctto.IdCecoCtto.CodCeco
            ws.cell(row=cont,column=14).value = ctto.IdCecoCtto.NomCeco
            ws.cell(row=cont,column=15).value = ctto.MonedaCtto
            ws.cell(row=cont,column=16).value = ctto.ValorCtto
            ws.cell(row=cont,column=17).value = ctto.AjusteCom
            ws.cell(row=cont,column=18).value = ctto.AjustNumEDP
            ws.cell(row=cont,column=19).value = ctto.AjustValEDP
            ws.cell(row=cont,column=20).value = ctto.AdjudicCtto
            ws.cell(row=cont,column=21).value = ctto.LocalCtto
            ws.cell(row=cont,column=22).value = ctto.TerrenCtto
            ws.cell(row=cont,column=23).value = ctto.SeguroCtto

            factor = fac(ctto.MonedaCtto)
            sumaODC = Odc.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('ValorODC'))['ValorODC__sum'] or 0
            sumaEDP = Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('ValEDP'))['ValEDP__sum'] or 0
            valcttoAct = ctto.ValorCtto + sumaODC
            commitment_ApProy =valcttoAct - ctto.AjusteCom
            auxiliar1 = ctto.ValorCtto
            auxiliar2 = ctto.AjusteCom
            edp_ApProy =sumaEDP - ctto.AjustValEDP
            commitment_togo =commitment_ApProy - edp_ApProy

            ws.cell(row=cont,column=24).value = sumaODC
            ws.cell(row=cont,column=25).value = sumaEDP
            ws.cell(row=cont,column=26).value = valcttoAct
            ws.cell(row=cont,column=27).value = commitment_ApProy
            ws.cell(row=cont,column=28).value = edp_ApProy
            ws.cell(row=cont,column=29).value = factor*edp_ApProy
            ws.cell(row=cont,column=30).value = factor*commitment_ApProy
            ws.cell(row=cont,column=31).value = factor*commitment_togo

            TerActualizado = (Odc.objects.filter(IdCtto__id=ctto.id).aggregate(Max('FechT_ODC'))['FechT_ODC__max']) or datetime(2009, 1, 1)
            #TerActualizado = datetime.strptime(TerActualizado, "%Y-%m-%d %H:%M:%S")


            if ctto.FechTerCtto.strftime('%F%H%M%S') > TerActualizado.strftime('%F%H%M%S'):
                TerActualizado = ctto.FechTerCtto

            ws.cell(row=cont,column=32).value = TerActualizado

            Fech_Sol_ultimaODC = (Odc.objects.filter(IdCtto__id=ctto.id).aggregate(Max('FechSolOdc'))['FechSolOdc__max']) or 0
            Fech_Apro_ultimaODC = (Odc.objects.filter(IdCtto__id=ctto.id).aggregate(Max('FechAppOdc'))['FechAppOdc__max']) or 0

            Fech_Pres_ultimaEDP = (Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Max('PresenEDP'))['PresenEDP__max']) or 0
            Fech_Apro_ultimaEDP = (Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Max('AprobEDP'))['AprobEDP__max']) or 0
            Fech_Period_ultimaEDP = (Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Max('PeriodEDPTer'))['PeriodEDPTer__max']) or 0


            ws.cell(row=cont,column=33).value = Fech_Sol_ultimaODC
            ws.cell(row=cont,column=34).value = Fech_Apro_ultimaODC

            ws.cell(row=cont,column=35).value = Fech_Pres_ultimaEDP
            ws.cell(row=cont,column=36).value = Fech_Apro_ultimaEDP
            ws.cell(row=cont,column=37).value = Fech_Period_ultimaEDP

            ws.cell(row=cont,column=38).value = ctto.FechSolCtto
            ws.cell(row=cont,column=39).value = ctto.FechAppCtto
            ws.cell(row=cont,column=40).value = ctto.IdCtta.RutCtta

            ws.cell(row=cont,column=41).value = ctto.ObservCtto

            ws.cell(row=cont,column=42).value = ctto.IdCtta.GiroCtta
            ws.cell(row=cont,column=43).value = ctto.IdCtta.DirCtta
            ws.cell(row=cont,column=44).value = ctto.IdCtta.ComunaCtta
            ws.cell(row=cont,column=45).value = ctto.IdCtta.CiudadCtta
            ws.cell(row=cont,column=46).value = ctto.ProvisCtto
            ws.cell(row=cont,column=47).value = factor*ctto.ValorCtto
            ws.cell(row=cont,column=48).value = factor*sumaODC
            ws.cell(row=cont,column=49).value = ctto.NumCtto+" "+ctto.DescCtto

            if  ctto.EstCtto == "":
                i = 0
            else:
                i = int(ctto.EstCtto)-1

            ws.cell(row=cont,column=50).value = Estatustxt[i]

            ws.cell(row=cont,column=51).value = ctto.CordCtto.Nombre
            ws.cell(row=cont,column=52).value = ctto.AdminCttoCtta.Nombre
            ws.cell(row=cont,column=53).value = ctto.AdminCttoCtta.Cargo
            ws.cell(row=cont,column=54).value = ctto.AdminCttoCtta.Correo
            ws.cell(row=cont,column=55).value = ctto.AdminCttoCtta.Cel

            sumaItemCtto= ItemCtto.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('TotalItem'))['TotalItem__sum'] or 0
            sumaItemOdc = ItemOdc.objects.filter(IdODC__IdCtto__id=ctto.id).aggregate(Sum('TotalItem'))['TotalItem__sum'] or 0
            sumaItemTotal = sumaItemCtto+sumaItemOdc

            ws.cell(row=cont,column=56).value = sumaItemCtto
            ws.cell(row=cont,column=57).value = sumaItemOdc
            ws.cell(row=cont,column=58).value = commitment_ApProy-(sumaItemTotal- ctto.AjusteCom)
            ws.cell(row=cont,column=59).value = sumaItemCtto*factor
            ws.cell(row=cont,column=60).value = sumaItemOdc*factor
            ws.cell(row=cont,column=61).value = sumaItemTotal*factor

            sumaRet = Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('RetEDP'))['RetEDP__sum'] or 0
            sumaDevRet = Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('DevRet'))['DevRet__sum'] or 0
            saldoRet = sumaRet - sumaDevRet

            ws.cell(row=cont,column=62).value = sumaRet
            ws.cell(row=cont,column=63).value = sumaDevRet
            ws.cell(row=cont,column=64).value = saldoRet*factor
            ws.cell(row=cont,column=65).value = factor


            cont = cont + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="ReportePersonasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        #response = HttpResponse(content_type="application/ms-excel")
        #contenido = "attachment; filename={0}".format(nombre_archivo)
        #response["Content-Disposition"] = contenido
        #wb.save(response)

        #['IdCtto', 'NumCtto', 'DescCtto', 'MonedaCtto', 'ValorCtto', 'IdCtta', 'EstCtto', 'FechIniCtto', 'FechTerCtto', 'IdCecoCtto', 'CordCtto', 'IdMandante',\
        #'Carpeta','TipoServ', 'AjusteCom','AjustNumEDP','AjustValEDP','AdjudicCtto','LocalCtto','TerrenCtto','SeguroCtto'],



        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'

        wb.save(response)
        return response

class ReporteCommitmentItem(TemplateView):

    #Usamos el metodo get para generar el archivo excel
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        CTTOS = Ctto.objects.all()
        ODC = Odc.objects.all()
        EDP = Edp.objects.all()
        ITEMCTTO = ItemCtto.objects.all()
        ITEMODC = ItemOdc.objects.all()

        Estatustxt = [ 'Solicitados en la semana','No iniciados','En preparación','En licitación','En evaluación','En negociación / Adjudicación','Adjudicado','En emisión','En firmas','En ejecución','Servicio terminado','En cierre','Cerrado','Servicio suspendido','Solicitud anulada','Solicitud diferida o postergada']


        wb = Workbook()

        ws = wb.active

        ws['A1'] = 'Centro Costo'
        ws['B1'] = 'Cuenta'
        ws['C1'] = 'Mandante'
        ws['D1'] = 'Tipo'
        ws['E1'] = 'N° Ctto'
        ws['F1'] = 'N° Odc'
        ws['G1'] = 'N° Item'
        ws['H1'] = 'Descripcion Servicio'
        ws['I1'] = 'Contratista'

        ws['J1'] = 'Estatus'
        ws['K1'] = 'Area'
        ws['L1'] = 'Cuenta'
        ws['M1'] = 'Descrip-Cuenta'
        ws['N1'] = 'Presupuesto'
        ws['O1'] = 'Moneda Ctto'
        ws['P1'] = 'Valor Item'
        ws['Q1'] = 'Valor Item (USD)'
        ws['R1'] = 'Ajuste Commitment'
        ws['S1'] = 'Ajuste Commitment (USD)'
        ws['T1'] = 'Valor Item Ajustado'
        ws['U1'] = 'Valor Item Ajustado (USD)'


        cont=2

        for Item in ITEMCTTO:

            factor = fac(Item.IdCtto.MonedaCtto)

            ws.cell(row=cont,column=1).value = Item.IdCtto.IdCecoCtto.IdAreas.CodArea
            ws.cell(row=cont,column=2).value = Item.IdCtto.IdCecoCtto.CodCeco
            ws.cell(row=cont,column=3).value = Item.IdCtto.IdMandante.NomMandte
            ws.cell(row=cont,column=4).value = Item.IdCtto.TipoServ
            ws.cell(row=cont,column=5).value = Item.IdCtto.NumCtto

            ws.cell(row=cont,column=6).value = 'ODC 0'
            ws.cell(row=cont,column=7).value = Item.NumItem

            ws.cell(row=cont,column=8).value = Item.IdCtto.DescCtto
            ws.cell(row=cont,column=9).value = Item.IdCtto.IdCtta.NomCtta

            ws.cell(row=cont,column=10).value = Item.IdCtto.EstCtto
            ws.cell(row=cont,column=11).value = Item.IdCtto.IdCecoCtto.IdAreas.NomArea
            ws.cell(row=cont,column=12).value = Item.IdCtto.IdCecoCtto.CodCeco
            ws.cell(row=cont,column=13).value = Item.IdCtto.IdCecoCtto.NomCeco
            ws.cell(row=cont,column=14).value = Item.PresupuestoItem.year
            ws.cell(row=cont,column=15).value = Item.IdCtto.MonedaCtto
            ws.cell(row=cont,column=16).value = Item.TotalItem
            ws.cell(row=cont,column=17).value = Item.TotalItem*factor

            #auxiliar1 = 0

            #if Item.NumItem == '1':
            #    Item.NumItem = "01"


            #if Item.NumItem == '01' or Item.NumItem == '1':
            #    ws.cell(row=cont,column=18).value = Item.IdCtto.AjusteCom
            #    ws.cell(row=cont,column=19).value = Item.IdCtto.AjusteCom*factor
            #    auxiliar1 = Item.IdCtto.AjusteCom

            #ItemAjustado = Item.TotalItem - auxiliar1
            #ws.cell(row=cont,column=20).value = ItemAjustado
            #ws.cell(row=cont,column=21).value = ItemAjustado*factor


            cont = cont + 1



        for Item in ITEMODC:

            factor = fac(Item.IdODC.IdCtto.MonedaCtto)

            ws.cell(row=cont,column=1).value = Item.IdODC.IdCtto.IdCecoCtto.IdAreas.CodArea
            ws.cell(row=cont,column=2).value = Item.IdODC.IdCtto.IdCecoCtto.CodCeco
            ws.cell(row=cont,column=3).value = Item.IdODC.IdCtto.IdMandante.NomMandte
            ws.cell(row=cont,column=4).value = Item.IdODC.IdCtto.TipoServ
            ws.cell(row=cont,column=5).value = Item.IdODC.IdCtto.NumCtto

            ws.cell(row=cont,column=6).value = Item.IdODC.NumODC
            ws.cell(row=cont,column=7).value = Item.NumItem

            ws.cell(row=cont,column=8).value = Item.IdODC.IdCtto.DescCtto
            ws.cell(row=cont,column=9).value = Item.IdODC.IdCtto.IdCtta.NomCtta

            ws.cell(row=cont,column=10).value = Item.IdODC.IdCtto.EstCtto
            ws.cell(row=cont,column=11).value = Item.IdODC.IdCtto.IdCecoCtto.IdAreas.NomArea
            ws.cell(row=cont,column=12).value = Item.IdODC.IdCtto.IdCecoCtto.CodCeco
            ws.cell(row=cont,column=13).value = Item.IdODC.IdCtto.IdCecoCtto.NomCeco
            ws.cell(row=cont,column=14).value = Item.PresupuestoItem.year
            ws.cell(row=cont,column=15).value = Item.IdODC.IdCtto.MonedaCtto
            ws.cell(row=cont,column=16).value = Item.TotalItem
            ws.cell(row=cont,column=17).value = Item.TotalItem*factor

            ws.cell(row=cont,column=18).value = 0
            ws.cell(row=cont,column=19).value = 0


            ItemAjustado = Item.TotalItem

            ws.cell(row=cont,column=20).value = ItemAjustado
            ws.cell(row=cont,column=21).value = factor*ItemAjustado

            cont = cont + 1



















        nombre_archivo ="ReportePersonasExcel.xlsx"


        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'

        wb.save(response)
        return response















class ReporteEDPExcel(TemplateView):

    #Usamos el metodo get para generar el archivo excel
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        CTTOS = Ctto.objects.all()
        ODC = Odc.objects.all()
        EDP = Edp.objects.all()

        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE EDP'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'Ctto'
        ws['C3'] = 'Ctta'
        ws['D3'] = 'Descripción'
        ws['E3'] = 'Nº EDP'
        ws['F3'] = 'Moneda'

        ws['G3'] = 'Valor EDP'
        ws['H3'] = 'Dev Anticipo'
        ws['I3'] = 'Reten EDP'
        ws['J3'] = 'Dev Ret EDP'

        ws['K3'] = 'Valor EDP [USD]'


        ws['L3'] = 'P inicio'
        ws['M3'] = 'P Termino'
        ws['N3'] = 'Fecha Present EDP'
        ws['O3'] = 'Fecha Aprob EDP'
        ws['P3'] = 'Obs EDP'
        ws['Q3'] = 'EstCtto'
        ws['R3'] = 'Ccosto'
        ws['S3'] = 'NomCeco'
        ws['T3'] = 'Mandante'

        cont=4
        valcttoAct = 0
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for ctto in CTTOS:
            factor = fac(ctto.MonedaCtto)
            for edp in Edp.objects.filter(IdCtto__id=ctto.id):

                ws.cell(row=cont,column=2).value = ctto.NumCtto
                ws.cell(row=cont,column=3).value = ctto.IdCtta.NomCtta
                ws.cell(row=cont,column=4).value = ctto.DescCtto
                ws.cell(row=cont,column=5).value = edp.NumEDP
                ws.cell(row=cont,column=6).value = ctto.MonedaCtto

                ws.cell(row=cont,column=7).value = edp.ValEDP
                ws.cell(row=cont,column=8).value = edp.DevAntEDP
                ws.cell(row=cont,column=9).value = edp.RetEDP
                ws.cell(row=cont,column=10).value = edp.DevRet

                ws.cell(row=cont,column=11).value = factor*edp.ValEDP

                ws.cell(row=cont,column=12).value = edp.PeriodEDP
                ws.cell(row=cont,column=13).value = edp.PeriodEDPTer
                ws.cell(row=cont,column=14).value = edp.PresenEDP
                ws.cell(row=cont,column=15).value = edp.AprobEDP
                ws.cell(row=cont,column=16).value = edp.ObservEDP
                ws.cell(row=cont,column=17).value = ctto.EstCtto
                ws.cell(row=cont,column=18).value = ctto.IdCecoCtto.CodCeco
                ws.cell(row=cont,column=19).value = ctto.IdCecoCtto.NomCeco
                ws.cell(row=cont,column=20).value = ctto.IdMandante.NomMandte


                cont = cont + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="ReportePersonasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        #response = HttpResponse(content_type="application/ms-excel")
        #contenido = "attachment; filename={0}".format(nombre_archivo)
        #response["Content-Disposition"] = contenido
        #wb.save(response)

        #['IdCtto', 'NumCtto', 'DescCtto', 'MonedaCtto', 'ValorCtto', 'IdCtta', 'EstCtto', 'FechIniCtto', 'FechTerCtto', 'IdCecoCtto', 'CordCtto', 'IdMandante',\
        #'Carpeta','TipoServ', 'AjusteCom','AjustNumEDP','AjustValEDP','AdjudicCtto','LocalCtto','TerrenCtto','SeguroCtto'],



        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=mydata_Edp.xlsx'

        wb.save(response)
        return response




class ReporteODCExcel(TemplateView):

    #Usamos el metodo get para generar el archivo excel
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        CTTOS = Ctto.objects.all()
        ODC = Odc.objects.all()
        EDP = Edp.objects.all()

        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE ODC'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'Ctto'
        ws['C3'] = 'Ctta'
        ws['D3'] = 'Descripción'
        ws['E3'] = 'Nº ODC'
        ws['F3'] = 'Valor ODC'
        ws['G3'] = 'Moneda'
        ws['H3'] = 'Valor ODC [USD]'
        ws['I3'] = 'Cuenta ODC'
        ws['J3'] = 'F Termino'
        ws['K3'] = 'Fecha Sol ODC'
        ws['L3'] = 'Fecha Aprob ODC'
        ws['M3'] = 'Obs ODC'
        ws['N3'] = 'EstCtto'

        cont=4
        valcttoAct = 0
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for ctto in CTTOS:
            factor = fac(ctto.MonedaCtto)
            for odc in Odc.objects.filter(IdCtto__id=ctto.id):

                ws.cell(row=cont,column=2).value = ctto.NumCtto
                ws.cell(row=cont,column=3).value = ctto.IdCtta.NomCtta
                ws.cell(row=cont,column=4).value = ctto.DescCtto
                ws.cell(row=cont,column=5).value = odc.NumODC
                ws.cell(row=cont,column=6).value = odc.ValorODC
                ws.cell(row=cont,column=7).value = ctto.MonedaCtto
                ws.cell(row=cont,column=8).value = factor*odc.ValorODC
                ws.cell(row=cont,column=9).value = odc.IdCecoODC.CodCeco
                ws.cell(row=cont,column=10).value = odc.FechT_ODC
                ws.cell(row=cont,column=11).value = odc.FechSolOdc
                ws.cell(row=cont,column=12).value = odc.FechAppOdc
                ws.cell(row=cont,column=13).value = odc.ObservOdc
                ws.cell(row=cont,column=14).value = ctto.EstCtto

                cont = cont + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="ReportePersonasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        #response = HttpResponse(content_type="application/ms-excel")
        #contenido = "attachment; filename={0}".format(nombre_archivo)
        #response["Content-Disposition"] = contenido
        #wb.save(response)

        #['IdCtto', 'NumCtto', 'DescCtto', 'MonedaCtto', 'ValorCtto', 'IdCtta', 'EstCtto', 'FechIniCtto', 'FechTerCtto', 'IdCecoCtto', 'CordCtto', 'IdMandante',\
        #'Carpeta','TipoServ', 'AjusteCom','AjustNumEDP','AjustValEDP','AdjudicCtto','LocalCtto','TerrenCtto','SeguroCtto'],



        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=mydata_Odc.xlsx'

        wb.save(response)
        return response


class ReporteITEMExcel(TemplateView):

    def get(self, request, *args, **kwargs):
            #Obtenemos todas las personas de nuestra base de datos

        CTTOS = Ctto.objects.all()
        ODC = Odc.objects.all()
        EDP = Edp.objects.all()
        ITEMCTTO = ItemCtto.objects.all()
        ITEMODC = ItemOdc.objects.all()

        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE ITEM'

        ws['B3'] = 'Ctto'
        ws['C3'] = 'Descripción'
        ws['D3'] = 'Ctta'
        ws['E3'] = 'NumItem'
        ws['F3'] = 'IdCecoCtto'
        ws['G3'] = 'DescripItem'
        ws['H3'] = 'UnidItem'
        ws['I3'] = 'CantItem'
        ws['J3'] = 'PuItem'
        ws['K3'] = 'TotalItem'
        ws['L3'] = 'ObservItem'
        ws['M3'] = 'PresupuestoItem'


        cont=4
        valcttoAct = 0
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for Item in ITEMCTTO:

                ws.cell(row=cont,column=2).value = Item.IdCtto.NumCtto
                ws.cell(row=cont,column=3).value = Item.IdCtto.DescCtto
                ws.cell(row=cont,column=4).value = Item.IdCtto.IdCtta.NomCtta
                ws.cell(row=cont,column=5).value = Item.NumItem
                ws.cell(row=cont,column=6).value = Item.IdCecoCtto.CodCeco
                ws.cell(row=cont,column=7).value = Item.DescripItem
                ws.cell(row=cont,column=8).value = Item.UnidItem
                ws.cell(row=cont,column=9).value = Item.CantItem
                ws.cell(row=cont,column=10).value = Item.PuItem
                ws.cell(row=cont,column=11).value = Item.TotalItem
                ws.cell(row=cont,column=12).value = Item.ObservItem
                ws.cell(row=cont,column=13).value = Item.PresupuestoItem

                if (Item.PresupuestoItem is None ):
                    Item.PresupuestoItem = date(Item.IdCtto.FechIniCtto.year,1,1)
                    Item.save()



                cont = cont + 1

        cont=300
        valcttoAct = 0

        for ctto in CTTOS:
                ws.cell(row=cont,column=2).value = ctto.NumCtto

                sumaItem = ItemCtto.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('TotalItem'))['TotalItem__sum'] or 0
                ws.cell(row=cont,column=3).value = sumaItem
                cuentaItem = ItemCtto.objects.filter(IdCtto__id=ctto.id).aggregate(Count('TotalItem'))['TotalItem__count'] or 0
                ws.cell(row=cont,column=4).value = cuentaItem
                ws.cell(row=cont,column=5).value = ctto.ValorCtto

                if cuentaItem == 0 :
                    p = ItemCtto(
                    NumItem = '01',
                    DescripItem = ctto.DescCtto,
                    UnidItem = 'Gl',
                    CantItem = 1,
                    PuItem = ctto.ValorCtto,
                    TotalItem = '',
                    ObservItem ='Item Automatico',
                    )

                    p.IdCecoCtto_id = ctto.IdCecoCtto_id
                    p.IdCtto_id =  ctto.id

                    p.save()


                cont = cont + 1



        ws['B599'] = 'REPORTE DE ITEM'
        ws['B600'] = 'Ctto'
        ws['C600'] = 'ODC'
        ws['D600'] = 'Ctta'
        ws['E600'] = 'NumItem'
        ws['F600'] = 'IdCecoCtto'
        ws['G600'] = 'DescripItem'
        ws['H600'] = 'UnidItem'
        ws['I600'] = 'CantItem'
        ws['J600'] = 'PuItem'
        ws['K600'] = 'TotalItem'
        ws['L600'] = 'ObservItem'
        ws['M600'] = 'PresupuestoItem'

        cont=605

        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for Item in ITEMODC:

                ws.cell(row=cont,column=2).value = Item.IdODC.IdCtto.NumCtto
                ws.cell(row=cont,column=3).value = Item.IdODC.NumODC
                ws.cell(row=cont,column=4).value = Item.IdODC.IdCtto.IdCtta.NomCtta
                ws.cell(row=cont,column=5).value = Item.NumItem
                ws.cell(row=cont,column=6).value = Item.IdCecoODC.CodCeco
                ws.cell(row=cont,column=7).value = Item.DescripItem
                ws.cell(row=cont,column=8).value = Item.UnidItem
                ws.cell(row=cont,column=9).value = Item.CantItem
                ws.cell(row=cont,column=10).value = Item.PuItem
                ws.cell(row=cont,column=11).value = Item.TotalItem
                ws.cell(row=cont,column=12).value = Item.ObservItem
                ws.cell(row=cont,column=13).value = Item.PresupuestoItem
                cont = cont + 1




        cont=800

        for odc in ODC:
                ws.cell(row=cont,column=2).value = odc.IdCtto.NumCtto
                ws.cell(row=cont,column=3).value = odc.NumODC
                sumaItem = ItemOdc.objects.filter(IdODC__id=odc.id).aggregate(Sum('TotalItem'))['TotalItem__sum'] or 0
                ws.cell(row=cont,column=4).value = sumaItem
                cuentaItem = ItemOdc.objects.filter(IdODC__id=odc.id).aggregate(Count('TotalItem'))['TotalItem__count'] or 0
                ws.cell(row=cont,column=5).value = cuentaItem
                ws.cell(row=cont,column=6).value = odc.ValorODC

                if cuentaItem == 0 :
                    p = ItemOdc(
                    NumItem = '01',
                    DescripItem = odc.DescripODC,
                    UnidItem = 'Gl',
                    CantItem = 1,
                    PuItem = odc.ValorODC,
                    TotalItem = '',
                    ObservItem ='Item Automatico',
                    )

                    p.IdCecoODC_id = odc.IdCecoODC_id
                    p.IdODC_id =  odc.id

                    p.save()



                cont = cont + 1



                cont = cont + 1
























        #Establecemos el nombre del archivo
        nombre_archivo ="ReportePersonasExcel.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=mydata_ItemCtto.xlsx'

        wb.save(response)
        return response




class ReporteFiniquito(TemplateView):

    #Usamos el metodo get para generar el archivo excel
    def get(self, request, *args, **kwargs):
            #Obtenemos todas las personas de nuestra base de datos

            CTTOS = Ctto.objects.all()
            ODC = Odc.objects.all()
            EDP = Edp.objects.all()

            #Creamos el libro de trabajo
            wb = Workbook()
            #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
            ws = wb.active


            wb = load_workbook(filename = 'DatosEDPCierre_Ctto_OS.xlsm', keep_vba=True)
            wb.template = False
            ws = wb.get_sheet_by_name('Datos')

            try:
                id_ctto = self.kwargs['id_ctto']


                ctto = Ctto.objects.get(id=id_ctto)
            except ObjectDoesNotExist:
                valor =0

            Desc_ceco = Ctto.objects.get(id=id_ctto).IdCecoCtto.CodCeco+': '+Ctto.objects.get(id=id_ctto).IdCecoCtto.NomCeco
            factor = fac(ctto.MonedaCtto)
            Item_ctto = ItemCtto.objects.filter(IdCtto__id=id_ctto).order_by('NumItem')
            Aportes_ctto = AportesCtto.objects.filter(IdCtto__id=id_ctto).order_by('NumItem')

            formato_fecha = "%Y-%m-%d"

            s_fhoy = fechaPalabra(time.strftime("%Y-%m-%d"))
            hoy = date.today()

            Finicio = Ctto.objects.get(id=self.kwargs['id_ctto']).FechIniCtto
            Ftermino = Ctto.objects.get(id=self.kwargs['id_ctto']).FechTerCtto
            Plazo = Plazodiaz(Finicio,Ftermino)
            PlazoPalabras = number_to_letter.to_word(int(Plazo))
            fecha_inicialPalabras =  fechaPalabra(Finicio)
            fecha_finalPalabras =  fechaPalabra(Ftermino)


            s_nommandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.NomMandte
            s_rutmandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.RutMandte
            s_direcmandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.DirecMandte
            s_comunmandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.ComunaMandte
            s_ciudmandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.CiudadMandte


            s_numctto = Ctto.objects.get(id=self.kwargs['id_ctto']).NumCtto
            s_descctto = Ctto.objects.get(id=self.kwargs['id_ctto']).DescCtto
            s_alcactto = Ctto.objects.get(id=self.kwargs['id_ctto']).AlcanceCtto
            s_monedctto = Ctto.objects.get(id=self.kwargs['id_ctto']).MonedaCtto
            s_valorctto = Ctto.objects.get(id=self.kwargs['id_ctto']).ValorCtto
            s_valorcttoUSD = Ctto.objects.get(id=self.kwargs['id_ctto']).ValorCtto*factor
            s_valorcttopalabras = number_to_letter.to_word(Ctto.objects.get(id=self.kwargs['id_ctto']).ValorCtto,str(ctto.MonedaCtto))
            s_modalidadctto = Ctto.objects.get(id=self.kwargs['id_ctto']).Modalidad
            s_ofertactto = Ctto.objects.get(id=self.kwargs['id_ctto']).DocOferta
            s_fechofertctto = fechaPalabra(Ctto.objects.get(id=self.kwargs['id_ctto']).FechOferta)
            s_iva = Ctto.objects.get(id=self.kwargs['id_ctto']).IvaOferta
            s_factor = factor


            s_nomctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.NomCtta
            s_rutctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.RutCtta
            s_dirctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.DirCtta
            s_comunctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.ComunaCtta
            s_ciudctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.CiudadCtta
            s_nomrep1ctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.Rep1Ctta
            s_rutrep1ctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.Rep1Ctta

            s_cordCtta = Ctto.objects.get(id=self.kwargs['id_ctto']).CordCtto.Nombre
            s_cargCordCtta = Ctto.objects.get(id=self.kwargs['id_ctto']).CordCtto.Cargo
            s_correoCordCtta = Ctto.objects.get(id=self.kwargs['id_ctto']).CordCtto.Correo

            s_coordctto = Ctto.objects.get(id=self.kwargs['id_ctto']).CordCtto
            s_nombccosto = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCecoCtto.NomCeco
            s_numccosto = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCecoCtto.CodCeco


            s_nomdueno = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCecoCtto.IdDueno.NomDueno
            s_rutdueno = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCecoCtto.IdDueno.RutDueno
            s_cargdueno = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCecoCtto.IdDueno.CargoDueno


            s_fechaperonariaMdte = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.FechDocpersonMandte
            s_fechaperonariaMdtepalabra = fechaPalabra(s_fechaperonariaMdte)
            s_notariaMdte = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.NotariapersonMandte



            s_nomcttocompleto =str(s_numctto)+" - '"+str(s_descctto)+"''"

            ValorBoleta = Ctto.objects.get(id=self.kwargs['id_ctto']).Boleta
            MonedaBoleta = Ctto.objects.get(id=self.kwargs['id_ctto']).MonedaBoleta
            VigenBoleta = Ctto.objects.get(id=self.kwargs['id_ctto']).VigenBoleta

            sumaODC = Odc.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('ValorODC'))['ValorODC__sum'] or 0
            sumaEDP = Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('ValEDP'))['ValEDP__sum'] or 0
            sumaReten = Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('RetEDP'))['RetEDP__sum'] or 0
            sumaDevRet = Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('DevRet'))['DevRet__sum'] or 0
            sumaAnticipo = Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('AnticipoEDP'))['AnticipoEDP__sum'] or 0
            sumaDevAnticipo = Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Sum('DevAntEDP'))['DevAntEDP__sum'] or 0
            sumaODCpalabras = number_to_letter.to_word(sumaODC)


            s_valcttoAct = s_valorctto + sumaODC
            s_valcttoActpalabras = number_to_letter.to_word(s_valcttoAct)
            s_saldoReten = sumaReten-sumaDevRet
            s_saldoAnticipo = sumaAnticipo-sumaDevAnticipo



            TerActualizado = (Odc.objects.filter(IdCtto__id=ctto.id).aggregate(Max('FechT_ODC'))['FechT_ODC__max']) or datetime(2009, 1, 1)
            #TerActualizado = datetime.strptime(TerActualizado, "%Y-%m-%d %H:%M:%S")


            if ctto.FechTerCtto.strftime('%F%H%M%S') > TerActualizado.strftime('%F%H%M%S'):
                TerActualizado = ctto.FechTerCtto

            Fech_ultPeriodEDP = (Edp.objects.filter(IdCtto__id=ctto.id).aggregate(Max('PeriodEDP'))['PeriodEDP__max']) or 0
            AuxUltEDP = Edp.objects.filter(IdCtto__id=ctto.id).latest('PeriodEDP')
            try:
                NumUltimoEDP = int(AuxUltEDP.NumEDP[-2:])
            except:
                NumUltimoEDP = 0

            ws['B3'] = s_nommandante
            ws['B7'] = s_numctto
            ws['B8'] = s_descctto
            ws['B9'] = s_alcactto
            ws['B10'] = s_numccosto+":"+s_nombccosto
            ws['B11'] = s_nomdueno
            ws['B12'] = s_cargdueno
            ws['B20'] = s_nomctta
            ws['B21'] = s_rutctta
            ws['B36'] = Finicio
            ws['B38'] = Ftermino
            ws['B42'] = s_monedctto
            ws['B43'] = s_valorctto
            ws['B45'] = s_iva
            ws['B55'] = s_valorcttoUSD
            ws['B63'] = s_cordCtta
            ws['B64'] = s_cargCordCtta
            ws['B78'] = s_factor
            ws['B85'] = NumUltimoEDP
            ws['B86'] = s_saldoReten
            ws['B87'] = s_saldoAnticipo
            ws['B88'] = sumaEDP
            ws['B89'] = sumaODC
            ws['B90'] = VigenBoleta
            ws['B91'] = TerActualizado


            cont = 2
            for odc in Odc.objects.filter(IdCtto__id=ctto.id):

                ws.cell(row=cont,column=29).value = odc.NumODC
                ws.cell(row=cont,column=30).value = odc.FechT_ODC
                ws.cell(row=cont,column=31).value = odc.IdCecoODC.CodCeco
                ws.cell(row=cont,column=32).value = odc.DescripODC
                ws.cell(row=cont,column=33).value = odc.ValorODC

                cont = cont + 1

            cont = 2
            for edp in Edp.objects.filter(IdCtto__id=ctto.id):

                ws.cell(row=cont,column=36).value = edp.NumEDP
                ws.cell(row=cont,column=37).value = edp.PeriodEDP
                ws.cell(row=cont,column=38).value = edp.PeriodEDPTer
                ws.cell(row=cont,column=39).value = edp.ValEDP
                ws.cell(row=cont,column=40).value = edp.DevAntEDP
                ws.cell(row=cont,column=41).value = edp.RetEDP
                ws.cell(row=cont,column=42).value = edp.DevRet
                ws.cell(row=cont,column=43).value = edp.FactEDP

                cont = cont + 1




            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Cierre_EDP_OS.xlsm'

            wb.save(response)



            tpl=DocxTemplate('test_files/BD_Finiquito.docx')

            sd = tpl.new_subdoc()

            context = {
                'Fhoy':s_fhoy,
                'NumCtto' : s_numctto,
                'Nom_Ctto' : s_descctto,
                'Alcan_Ctto' : s_alcactto,
                'Nom_Mdte' : s_nommandante,
                'Rut_Mdte': s_rutmandante,
                'Direcc_Mdte' :s_direcmandante,
                'Comu_Mdte' : s_comunmandante,
                'Ciud_Mdte' : s_ciudmandante,
                'Nom_RepLeg1_NU' : s_nomdueno,
                'Rut_RepLeg1_NU' : s_rutdueno,
                'Carg_RepLeg1_NU' : s_cargdueno,
                'Nom_RepLeg2_NU' : 'Julio Retamal',
                'Rut_RepLeg2_NU': '9.727.084-8',
                'Carg_RepLeg2_NU' : '',
                'Nom_Ctta' : s_nomctta,
                'Rut_Ctta' : s_rutctta,
                'Direcc_Ctta' : s_dirctta,
                'Comu_Ctta' : s_comunctta,
                'Ciud_Ctta' : s_ciudctta,
                'Nom_RepLeg1_Ctta' : s_nomrep1ctta,
                'Rut_RepLeg1_Ctta' : s_rutrep1ctta,
                'Moneda_Serv' : s_monedctto,
                'Valor_Serv' : s_valorctto,
                'Valor_Serv_Palabras' : s_valorcttopalabras,
                'Mod_Servicio' :s_modalidadctto,
                'Dur_Serv' : Plazo,
                'Dur_Serv_Palabras' : PlazoPalabras,
                'Fecha_IniServ_Palabras' :fecha_inicialPalabras,
                'Fecha_Ter_Serv_Palabras':fecha_finalPalabras,
                'Documento_Oferta_Ctta':s_ofertactto,
                'Fecha_Oferta_Ctta' :s_fechofertctto,

                'Nom_Coord_Mdte' :s_cordCtta,
                'Cargo_Coord_Mdte':s_cargCordCtta,
                'Correo_Coord_Mdte' :s_correoCordCtta,
                'Valor_SumOdc':sumaODC,
                'Valor_SumOdc_Palabras':sumaODCpalabras,
                'ValorAct_Serv':s_valcttoAct,
                'ValorAct_Serv_Palabras':s_valcttoActpalabras,

                'fechapersoneriapalabra':s_fechaperonariaMdtepalabra,
                'notariamandente':s_notariaMdte,



                'date' : '2016-03-17',
                'example' : '',
            }

            tpl.render(context)

            nombrArchivo='test_files/Finiquito_'+s_numctto+'.docx'
            tpl.save(nombrArchivo)


            return response











class crear_docCtto(TemplateView):
    def get(self, request, *args, **kwargs):

        wb = load_workbook(filename = 'DatosOS_Cttos.xlsm', keep_vba=True)
        wb.template = False
        ws = wb.get_sheet_by_name('Datos')

        try:
            id_ctto = self.kwargs['id_ctto']


            ctto = Ctto.objects.get(id=id_ctto)
        except ObjectDoesNotExist:
            valor =0

        Desc_ceco = Ctto.objects.get(id=id_ctto).IdCecoCtto.CodCeco+': '+Ctto.objects.get(id=id_ctto).IdCecoCtto.NomCeco
        factor = fac(ctto.MonedaCtto)
        Item_ctto = ItemCtto.objects.filter(IdCtto__id=id_ctto).order_by('NumItem')
        Aportes_ctto = AportesCtto.objects.filter(IdCtto__id=id_ctto).order_by('NumItem')

        formato_fecha = "%Y-%m-%d"

        s_fhoy = fechaPalabra(time.strftime("%Y-%m-%d"))
        hoy = date.today()
        devolcartaadj = hoy + timedelta(days=7)
        print ("devolver carta")
        print(devolcartaadj)

        s_devolcartaadj =fechaPalabra(devolcartaadj)

        Finicio = Ctto.objects.get(id=self.kwargs['id_ctto']).FechIniCtto
        Ftermino = Ctto.objects.get(id=self.kwargs['id_ctto']).FechTerCtto
        Plazo = Plazodiaz(Finicio,Ftermino)
        PlazoPalabras = number_to_letter.to_word(int(Plazo))
        fecha_inicialPalabras =  fechaPalabra(Finicio)
        fecha_finalPalabras =  fechaPalabra(Ftermino)


        s_nommandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.NomMandte
        s_rutmandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.RutMandte
        s_direcmandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.DirecMandte
        s_comunmandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.ComunaMandte
        s_ciudmandante = Ctto.objects.get(id=self.kwargs['id_ctto']).IdMandante.CiudadMandte


        s_numctto = Ctto.objects.get(id=self.kwargs['id_ctto']).NumCtto
        s_descctto = Ctto.objects.get(id=self.kwargs['id_ctto']).DescCtto
        s_alcactto = Ctto.objects.get(id=self.kwargs['id_ctto']).AlcanceCtto
        s_monedctto = Ctto.objects.get(id=self.kwargs['id_ctto']).MonedaCtto
        s_valorctto = Ctto.objects.get(id=self.kwargs['id_ctto']).ValorCtto
        s_valorcttoUSD = Ctto.objects.get(id=self.kwargs['id_ctto']).ValorCtto*factor
        s_valorcttopalabras = number_to_letter.to_word(Ctto.objects.get(id=self.kwargs['id_ctto']).ValorCtto,str(ctto.MonedaCtto))
        s_modalidadctto = Ctto.objects.get(id=self.kwargs['id_ctto']).Modalidad
        s_ofertactto = Ctto.objects.get(id=self.kwargs['id_ctto']).DocOferta
        s_fechofertctto = fechaPalabra(Ctto.objects.get(id=self.kwargs['id_ctto']).FechOferta)


        s_nomctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.NomCtta
        s_rutctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.RutCtta
        s_dirctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.DirCtta
        s_comunctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.ComunaCtta
        s_ciudctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.CiudadCtta
        s_nomrep1ctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.Rep1Ctta
        s_rutrep1ctta = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCtta.Rep1Ctta

        s_cordCtta = Ctto.objects.get(id=self.kwargs['id_ctto']).CordCtto.Nombre
        s_cargCordCtta = Ctto.objects.get(id=self.kwargs['id_ctto']).CordCtto.Cargo
        s_correoCordCtta = Ctto.objects.get(id=self.kwargs['id_ctto']).CordCtto.Correo

        s_numccosto = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCecoCtto.CodCeco
        s_nombccosto = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCecoCtto.NomCeco


        s_nomdueno = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCecoCtto.IdDueno.NomDueno
        s_cargdueno = Ctto.objects.get(id=self.kwargs['id_ctto']).IdCecoCtto.IdDueno.CargoDueno

        s_nomcttocompleto =str(s_numctto)+" - '"+str(s_descctto)+"''"

        ValorBoleta = Ctto.objects.get(id=self.kwargs['id_ctto']).Boleta
        MonedaBoleta = Ctto.objects.get(id=self.kwargs['id_ctto']).MonedaBoleta

        s_col_etiqaportes = ['Aporte del Mandante']
        s_aportesmdte  =[]
        for aporte in Aportes_ctto:
            s_aportesmdte.append({'label':aporte.NumItem,'cols':[aporte.Aporte]})

        s_col_etiqitem = ['Cuenta','Descrip','unid','cant','PU','Total']
        s_itemctto  =[]
        for item in Item_ctto:
            s_itemctto.append({'label':item.NumItem,'cols':[item.IdCecoCtto.CodCeco,item.DescripItem,item.UnidItem,item.CantItem,item.PuItem,item.TotalItem]})

        print(s_itemctto)
        s_itemctto.append({'label':'','cols':['','','','','Total('+s_monedctto+'):',s_valorctto]})

        if ValorBoleta != None and ValorBoleta !=0:
            try:

                ValorboletaPalabras =number_to_letter.to_word(Ctto.objects.get(id=self.kwargs['id_ctto']).Boleta,MonedaBoleta)
                FVigenciaBoleta = Ctto.objects.get(id=self.kwargs['id_ctto']).FechVigenBoleta
                FVigenciaBoletaPalabras = fechaPalabra(FVigenciaBoleta)
                print(ValorboletaPalabras)
                print(Finicio)
                print(FVigenciaBoleta)
                print(fecha_inicialPalabras)
                print(FVigenciaBoletaPalabras)
                print (locale.format("%d",ValorBoleta, grouping=True))
                print ('{:0,.2f}'.format(ValorBoleta))
                locale.setlocale( locale.LC_ALL, '' )
                print(locale.currency( ValorBoleta, grouping = True ))

                rt = RichText('an exemple of ')
                rt.add('El Contratista deberá entregar a El Mandante, en un plazo no superior a 30 días hábiles posteriores a la emisión de esta Carta de Adjudicación, una boleta de garantía bancaria ', style='Estilo1')
                rt.add(' a la vista e incondicional por el fiel cumplimiento del Contrato, emitida por un banco comercial autorizado para operar en el país,por un total de Dicha boleta, en su glosa')
                rt.add(', deberá indicar que su objeto es garantizar el fiel cumplimiento del Contrato N° El período de vigencia de dicha boleta, abarcará toda la duración del servicio', italic=True)
                rt.add(', hasta los 90 días siguientes a partir de la fecha fijada como término del Contrato')
                rt.add('some violet', color='#ff00ff')
                rt.add(' and ')
                rt.add('some striked', strike=True)
                rt.add(' and ')
                rt.add('some small', size=14)


                p1_Boleta = "El Contratista deberá entregar a El Mandante, en un plazo no superior a 30 días hábiles posteriores " +\
                                "a la emisión de esta Carta de Adjudicación, una boleta de garantía bancaria a la vista e incondicional "

                p2_Boleta ="por el fiel cumplimiento del Contrato, emitida por un banco comercial autorizado para operar en el país, " +\
                                "a favor de " + s_nommandante +" , R.U.T.: "+ s_rutmandante+" , por un total de "+str(MonedaBoleta)+" "+locale.format("%d",ValorBoleta, grouping=True)+" ( "+ ValorboletaPalabras+" )"

                p3_Boleta = " Dicha boleta, en su glosa, deberá indicar que su objeto es garantizar el fiel cumplimiento del Contrato N° " +\
                                 s_nomcttocompleto + ". El período de vigencia de dicha boleta, abarcará toda la duración del "

                p4_Boleta ="servicio, hasta los 90 días siguientes a partir de la fecha fijada como término del Contrato, esto es hasta el día "+\
                                 FVigenciaBoletaPalabras + ". "

                s_aplicaboleta = ""

                s_parrafoboleta = p1_Boleta+p2_Boleta+p3_Boleta+p4_Boleta



            except:
                #pdb.set_trace() ## Punto de ruptura
                s_parrafoboleta = "Exept"
                FVigenciaBoletaPalabras = "Exept"
                s_aplicaboleta = "Exept"
        else:
            s_parrafoboleta = ""
            FVigenciaBoletaPalabras = ""
            s_aplicaboleta = " (No Aplica Contrato " + s_numctto + " )"






        ws['B7'] = s_numctto
        ws['B8'] = s_descctto
        ws['B9'] = s_alcactto
        ws['B10'] = s_numccosto+":"+s_nombccosto
        ws['B11'] = s_nomdueno
        ws['B12'] = s_cargdueno
        ws['B20'] = s_nomctta
        ws['B21'] = s_rutctta
        ws['B36'] = Finicio
        ws['B38'] = Ftermino
        ws['B42'] = s_monedctto
        ws['B55'] = s_valorcttoUSD






        cont =1
        for item in Item_ctto:
            ws.cell(row=cont+1,column=7).value = item.NumItem
            ws.cell(row=cont+1,column=8).value = item.IdCecoCtto.CodCeco
            ws.cell(row=cont+1,column=9).value = item.DescripItem
            ws.cell(row=cont+1,column=10).value = item.UnidItem
            ws.cell(row=cont+1,column=11).value = item.CantItem
            ws.cell(row=cont+1,column=12).value = item.PuItem
            ws.cell(row=cont+1,column=13).value = item.TotalItem
            cont =cont+1

        #wb.save('DatosOS_Cttos.xlsx')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=DatosOS_Cttos.xlsm'

        wb.save(response)



        #tpl=DocxTemplate('test_files/header_footer_tpl.docx')
        tpl=DocxTemplate('test_files/Carta AdjudicacionBD.docx')

        sd = tpl.new_subdoc()
        p = sd.add_paragraph('El Contratista deberá entregar a El Mandante, en un plazo no superior a 30 días hábiles posteriores a la emisión de esta Carta de Adjudicación, una boleta de garantía bancaria a la vista e incondicional por el fiel cumplimiento del Contrato, emitida por un banco comercial autorizado para operar en el país,por un total de Dicha boleta, en su glosa, deberá indicar que su objeto es garantizar el fiel cumplimiento del Contrato N° El período de vigencia de dicha boleta, abarcará toda la duración del servicio, hasta los 90 días siguientes a partir de la fecha fijada como término del Contrato, esto es hasta el día y noche' )

        context = {
            'Fhoy':s_fhoy,
            'NumCtto' : s_numctto,
            'Nom_Ctto' : s_descctto,
            'Alcan_Ctto' : s_alcactto,
            'Nom_Mdte' : s_nommandante,
            'Rut_Mdte': s_rutmandante,
            'Direcc_Mdte' :s_direcmandante,
            'Comu_Mdte' : s_comunmandante,
            'Ciud_Mdte' : s_ciudctta,
            'Nom_RepLeg1_NU' : s_nomdueno,
            'Carg_RepLeg1_NU' : s_cargdueno,
            'Nom_RepLeg2_NU' : '',
            'Carg_RepLeg2_NU' : '',
            'Nom_Ctta' : s_nomctta,
            'Rut_Ctta' : s_rutctta,
            'Direcc_Ctta' : s_dirctta,
            'Comu_Ctta' : s_comunctta,
            'Ciud_Ctta' : s_ciudctta,
            'Nom_RepLeg1_Ctta' : s_nomrep1ctta,
            'Moneda_Serv' : s_monedctto,
            'Valor_Serv' : s_valorctto,
            'Valor_Serv_Palabras' : s_valorcttopalabras,
            'Mod_Servicio' :s_modalidadctto,
            'Dur_Serv' : Plazo,
            'Dur_Serv_Palabras' : PlazoPalabras,
            'Fecha_IniServ_Palabras' :fecha_inicialPalabras,
            'Fecha_Ter_Serv_Palabras':fecha_finalPalabras,
            'Documento_Oferta_Ctta':s_ofertactto,
            'Fecha_Oferta_Ctta' :s_fechofertctto,
            'Parrafo_Boleta':s_parrafoboleta,
            'AplicaBoleta':s_aplicaboleta,
            'col_etiqaportes':s_col_etiqaportes,
            'Tbl_aportes': s_aportesmdte,
            'col_etiqitem':s_col_etiqitem,
            'Tbl_itemc': s_itemctto,
            'Fecha_devolcarta':s_devolcartaadj,
            'Nom_Coord_Mdte' :s_cordCtta,
            'Cargo_Coord_Mdte':s_cargCordCtta,
            'Correo_Coord_Mdte' :s_correoCordCtta,




            'date' : '2016-03-17',
            'example' : '',
        }

        tpl.render(context)

        nombrArchivo='test_files/CartaAdj_'+s_numctto+'.docx'
        tpl.save(nombrArchivo)


        return response





class crear_docODC(TemplateView):
    def get(self, request, *args, **kwargs):



        wb = load_workbook(filename = 'Datos.xlsx')
        wb.template = False
        ws = wb.get_sheet_by_name('BD')

        try:
            id_odc = self.kwargs['id_odc']
            id_ctto = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.id

            ctto = Ctto.objects.get(id=id_ctto)
        except ObjectDoesNotExist:
            valor =0

        Desc_ceco = Odc.objects.get(id=id_odc).IdCecoODC.CodCeco+': '+Odc.objects.get(id=id_odc).IdCecoODC.NomCeco
        factor = fac(ctto.MonedaCtto)
        ODC_ctto = Odc.objects.filter(IdCtto__id=id_ctto).order_by('NumODC')
        Item_odc = ItemOdc.objects.filter(IdODC__id=id_odc).order_by('NumItem')
        TerActualizado_ant = ctto.FechTerCtto
        sumaODC = 0
        for odc in ODC_ctto:
            if odc.id == int(id_odc):
                print ("ODC Actual : " + odc.NumODC)
                break
            print("Fecha TAct")
            print(odc.FechT_ODC)
            if odc.FechT_ODC != None:
                TerActualizado_ant = odc.FechT_ODC
            sumaODC = sumaODC+ (odc.ValorODC or 0)


        ws['B5'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.NumCtto
        ws['B6'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.DescCtto
        ws['B7'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.IdCtta.NomCtta
        ws['B8'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.IdCtta.RutCtta
        ws['B9'] = Odc.objects.get(id=self.kwargs['id_odc']).NumODC
        ws['B10'] = ""
        ws['B11'] = ""
        ws['B12'] = Odc.objects.get(id=self.kwargs['id_odc']).FechAppOdc
        ws['B13'] = Odc.objects.get(id=self.kwargs['id_odc']).DescripODC
        ws['B14'] = Desc_ceco
        ws['B15'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCecoODC.IdDueno.NomDueno
        ws['B16'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCecoODC.IdDueno.CargoDueno
        ws['B17'] = ""
        ws['B18'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.MonedaCtto
        ws['B19'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.ValorCtto
        ws['B20'] = sumaODC
        ws['B21'] = Odc.objects.get(id=self.kwargs['id_odc']).ValorODC
        ws['B22'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.FechIniCtto
        ws['B23'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.FechTerCtto
        ws['B24'] = TerActualizado_ant
        ws['B25'] = Odc.objects.get(id=self.kwargs['id_odc']).FechT_ODC
        ws['B26'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.IdCtta.DirCtta
        ws['B27'] = ""
        ws['B28'] = ""
        ws['B29'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.IdMandante.NomMandte



        # Valores de ODC en USD
        ws['C19'] = Odc.objects.get(id=self.kwargs['id_odc']).IdCtto.ValorCtto*factor
        ws['C20'] = sumaODC*factor
        ws['C21'] = Odc.objects.get(id=self.kwargs['id_odc']).ValorODC*factor



        cont =1
        for itemodc in Item_odc:
            ws.cell(row=cont+30,column=1).value = itemodc.NumItem
            ws.cell(row=cont+30,column=2).value = itemodc.IdCecoODC.CodCeco
            ws.cell(row=cont+30,column=3).value = itemodc.DescripItem
            ws.cell(row=cont+30,column=4).value = itemodc.UnidItem
            ws.cell(row=cont+30,column=5).value = itemodc.CantItem
            ws.cell(row=cont+30,column=6).value = itemodc.PuItem
            ws.cell(row=cont+30,column=7).value = itemodc.TotalItem
            cont =cont+1

        #wb.save('Datos.xlsx')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Datos.xlsx'

        wb.save(response)

        return response









class crear_docEDP(TemplateView):
    def get(self, request, *args, **kwargs):

        wb = load_workbook(filename = 'CaratulaEDP.xlsx')
        wb.template = False
        ws = wb.get_sheet_by_name('BD')

        try:
            id_edp = self.kwargs['pk']
            id_ctto = Edp.objects.get(id=self.kwargs['pk']).IdCtto.id

            ctto = Ctto.objects.get(id=id_ctto)
        except ObjectDoesNotExist:
            valor =0

        factor = fac(ctto.MonedaCtto)
        EDP_ctto = Edp.objects.filter(IdCtto__id=id_ctto).order_by('NumEDP')
        ODC_ctto = Odc.objects.filter(IdCtto__id=id_ctto).order_by('NumODC')

        #Item_odc = ItemOdc.objects.filter(IdEDP__id=id_edp).order_by('NumItem')
        TerActualizado_ant = ctto.FechTerCtto


        sumaODC = 0
        for odc in ODC_ctto:
            if odc.id == int(id_odc):
                print ("ODC Actual : " + odc.NumODC)
                break
            print("Fecha TAct")
            print(odc.FechT_ODC)
            if odc.FechT_ODC != None:
                TerActualizado_ant = odc.FechT_ODC
            sumaODC = sumaODC+ (odc.ValorODC or 0)


        sumaEDP = 0
        sumaRet = 0
        SumaDevRet = 0

        for edp in EDP_ctto:
            sumaEDP = sumaEDP+ (edp.ValEDP or 0)
            sumaRet = sumaRet+ (edp.RetEDP or 0)
            SumaDevRet = SumaDevRet + (edp.DevRet or 0)

        ws['B5'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.NumCtto
        ws['B6'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.DescCtto
        ws['B7'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.IdCtta.NomCtta
        ws['B8'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.IdCtta.RutCtta
        ws['B9'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.IdMandante.NomMandte
        ws['B10'] = Edp.objects.get(id=self.kwargs['pk']).PeriodEDP
        ws['B11'] = Edp.objects.get(id=self.kwargs['pk']).PeriodEDPTer
        ws['B12'] = ""
        ws['B13'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.FechIniCtto
        ws['B14'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.FechTerCtto
        ws['B15'] = ""
        ws['B16'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.ValorCtto
        ws['B17'] = sumaODC
        ws['B18'] = Edp.objects.get(id=self.kwargs['pk']).NumEDP
        ws['B19'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.MonedaCtto
        ws['B20'] = ""
        ws['B21'] = ""
        ws['B22'] = ""
        ws['B23'] = Edp.objects.get(id=self.kwargs['pk']).RetEDP
        ws['B24'] = sumaRet-SumaDevRet
        ws['B25'] = Edp.objects.get(id=self.kwargs['pk']).ValEDP
        ws['B26'] = Edp.objects.get(id=self.kwargs['pk']).ValEDP*factor
        ws['B27'] = sumaEDP
        ws['B28'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.IdCecoCtto.CodCeco+': '+Edp.objects.get(id=self.kwargs['pk']).IdCtto.IdCecoCtto.NomCeco
        ws['B29'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.IdCecoCtto.IdDueno.NomDueno
        ws['B30'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.IvaOferta
        ws['B31'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.RetenCtto
        ws['B32'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.FechVigenBoleta
        ws['B33'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.TerrenCtto
        # Valores de ODC en USD




        #cont =1
        #for itemodc in Item_odc:
        #    ws.cell(row=cont+30,column=1).value = itemodc.NumItem
        #    ws.cell(row=cont+30,column=2).value = itemodc.IdCecoODC.CodCeco
        #    ws.cell(row=cont+30,column=3).value = itemodc.DescripItem
        #    ws.cell(row=cont+30,column=4).value = itemodc.UnidItem
        #    ws.cell(row=cont+30,column=5).value = itemodc.CantItem
        #    ws.cell(row=cont+30,column=6).value = itemodc.PuItem
        #    ws.cell(row=cont+30,column=7).value = itemodc.TotalItem
        #    cont =cont+1

        #wb.save('Datos.xlsx')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=CaratulaEDP.xlsx'

        wb.save(response)

        return response


























class Bienvenida(TemplateView):
    #template_name = 'Tabla_Servicios.html'
    template_name = 'form.html'
# Create your views here.




class CrearContrato(CreateView):
    model = Ctto
    #fields =['dni','nombre','apellido_paterno','apellido_materno']
    template_name = 'crear_contrato_new.html'
    form_class = CttoUpdateForm
    success_url = reverse_lazy('personas:personas')

    # Primer Formset
    def get_context_data(self, **kwargs):
        data = super(CrearContrato, self).get_context_data(**kwargs)

        if self.request.POST:
             data['ItemCttos'] = ItemCttoFormSet(self.request.POST)
             data['AportesCttos'] = AportesCttoFormSet(self.request.POST)
             data['MultasPcCttos'] = MultasPerClaveCttoFormSet(self.request.POST)
        else:
             data['ItemCttos'] = ItemCttoFormSet()
             data['AportesCttos'] = AportesCttoFormSet()
             data['MultasPcCttos'] = MultasPerClaveCttoFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        ItemCttos = context['ItemCttos']
        AportesCttos = context['AportesCttos']
        MultasPcCttos = context['MultasPcCttos']
        with transaction.atomic():
            self.object = form.save()

            if ItemCttos.is_valid():
                ItemCttos.instance = self.object
                ItemCttos.save()

            if AportesCttos.is_valid():
                AportesCttos.instance = self.object
                AportesCttos.save()

            if MultasPcCttos.is_valid():
                MultasPcCttos.instance = self.object
                MultasPcCttos.save()


        return super(CrearContrato, self).form_valid(form)




class ModificarContrato(UpdateView):
    #Especificamos que el modelo a utilizar va a ser Ctto
    form_class = CttoUpdateForm

    model = Ctto
    #Establecemos que la plantilla se llamara modificar persona
    template_name = 'modificar_contrato_new.html'
    #Determinamos los campos con los que se va a trabajar, esto es obligatorio sino nos saldra un error
    #fields = ['NumCtto','DescCtto','MonedaCtto','ValorCtto','IdCtta','EstCtto','FechIniCtto','IdCecoCtto','CordCtto','IdMandante' ]
    #Con esta linea establecemos que se hara despues que la operacion de modificacion se complete correctamente
    success_url = reverse_lazy('personas:personas')

    def get_context_data(self, **kwargs):
        data = super(ModificarContrato, self).get_context_data(**kwargs)
        dato = self.kwargs['pk']
        print('dato')
        print(dato)

        if self.request.POST:
            data['ItemCttos'] = ItemCttoFormSet(self.request.POST, instance=self.object)
            data['AportesCttos'] = AportesCttoFormSet(self.request.POST, instance=self.object)
            data['MultasPcCttos'] = MultasPerClaveCttoFormSet(self.request.POST, instance=self.object)
        else:
            data['ItemCttos'] = ItemCttoFormSet(instance=self.object)
            data['AportesCttos'] = AportesCttoFormSet(instance=self.object)
            data['MultasPcCttos'] = MultasPerClaveCttoFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        ItemCttos = context['ItemCttos']
        AportesCttos = context['AportesCttos']
        MultasPcCttos = context['MultasPcCttos']
        with transaction.atomic():
            self.object = form.save()

            if ItemCttos.is_valid():
                ItemCttos.instance = self.object
                ItemCttos.save()

            if AportesCttos.is_valid():
                AportesCttos.instance = self.object
                AportesCttos.save()

            if MultasPcCttos.is_valid():
                MultasPcCttos.instance = self.object
                MultasPcCttos.save()

        return super(ModificarContrato, self).form_valid(form)





class DetalleContrato(DetailView):
    model = Ctto
    template_name = 'detalle_persona_new.html'


#def sumarLista(lista):
#    sum=0
#    for i in range(0,len(lista)):
#        sum=sum+lista[i]
#    return sum






class Personas(ListView):
    model = Ctto
    template_name = 'personas_new.html'
    context_object_name = 'Cttos'
    buscar = "Contrato"

    def get_context_data(self, **kwargs):
        # Llamamos ala implementacion primero del  context
        context = super(Personas, self).get_context_data(**kwargs)
        # Agregamos el publisher
        context['buscar'] = self.buscar
        return context

class EditarPersona(UpdateView):
    model = Persona

    #fields =['dni','nombre','apellido_paterno','apellido_materno']
    form_class = PersonaCreateForm
    template_model = 'crear_persona.html'
    #success_url = reverse_lazy('personas:personas')

    # def get(self,request,*arg,**kwargs):
    #    id_edit = request.GET['id']

class ficha(ListView):
    model = Ctto
    template_name = 'ficha.html'
    context_object_name = 'Cttos'



def EditarContrato(request,id_ctto):
    CTTO = Ctto.objects.all()
    ODC = Odc.objects.all()
    EDP = Edp.objects.all()

    valor = 0
    try:
        valor = Ctto.objects.get(NumCtto=id_ctto).id
        ctto = Ctto.objects.get(NumCtto=id_ctto)
    except ObjectDoesNotExist:
        valor =0

    sumaODC = Odc.objects.filter(IdCtto__id=valor).aggregate(Sum('ValorODC'))['ValorODC__sum'] or 0
    sumaEDP = Edp.objects.filter(IdCtto__id=valor).aggregate(Sum('ValEDP'))['ValEDP__sum'] or 0
    sumaAnt = Edp.objects.filter(IdCtto__id=valor).aggregate(Sum('AnticipoEDP'))['AnticipoEDP__sum'] or 0
    sumaDevAnt = Edp.objects.filter(IdCtto__id=valor).aggregate(Sum('DevAntEDP'))['DevAntEDP__sum'] or 0
    sumaRet = Edp.objects.filter(IdCtto__id=valor).aggregate(Sum('RetEDP'))['RetEDP__sum'] or 0
    sumaDevRet = Edp.objects.filter(IdCtto__id=valor).aggregate(Sum('DevRet'))['DevRet__sum'] or 0
    sumaDesc = Edp.objects.filter(IdCtto__id=valor).aggregate(Sum('DescuentoEDP'))['DescuentoEDP__sum'] or 0

    TerActualizado = (Odc.objects.filter(IdCtto__id=valor).aggregate(Max('FechT_ODC'))['FechT_ODC__max']) or datetime(2009, 1, 1)
    if ctto.FechTerCtto.strftime('%F%H%M%S') > TerActualizado.strftime('%F%H%M%S'):
        TerActualizado = ctto.FechTerCtto


    ValActCtto = ctto.ValorCtto + sumaODC
    ValActFechTermCtto = TerActualizado
    Saldocontrato =ValActCtto-sumaEDP

    # Fecha ingresada
    #fecha_ingresada = '09/04/2008'

    # Separo el formato de fecha para convertirlo en yyyy/mm/dd
    #d = fecha_ingresada.split('/')
    #fecha_a_calcular = datetime.strptime(d[2] + d[1] + d[0],'%Y%m%d').date()

    # Creo una variable con la operación aritmética
    PlazoCtto = TerActualizado - ctto.FechIniCtto
    VigenciaCtto = TerActualizado - datetime.now().date()

    try:
        porvigencia =round(VigenciaCtto/PlazoCtto,1)
    except:
        porvigencia =0

    try:
        poravance =round(sumaEDP/ValActCtto,1)
    except:
        poravance =0


    return render_to_response('editar_contratos_new.html',{'Ctto':ctto,'Odc':ODC,'Edp':EDP,'id_ctto':valor,\
    'ValActCtto':ValActCtto,'TerActualizado':TerActualizado,'sumaODC':sumaODC,'sumaEDP':sumaEDP,'sumaAnt':sumaAnt,\
    'sumaDevAnt':sumaDevAnt,'sumaRet':sumaRet,'sumaDevRet': sumaDevRet,'sumaDesc':sumaDesc,'Vigencia':VigenciaCtto,\
    'porvigencia':porvigencia*100,'poravance':poravance*100,'saldocontrato': Saldocontrato
     })






class DetalleEdp(DetailView):
    model = Edp
    template_name = 'detalle_Edp_new.html'


# def sumarLista(lista):
#    sum=0
#    for i in range(0,len(lista)):
#        sum=sum+lista[i]
#    return sum

class ModificarEdp(UpdateView):
    #Especificamos que el modelo a utilizar va a ser Persona
    form_class = EdpUpdateForm

    model = Edp
    #Establecemos que la plantilla se llamara modificar persona
    template_name = 'modificar_edp_new.html'
    #Determinamos los campos con los que se va a trabajar, esto es obligatorio sino nos saldra un error
    #fields = ['NumCtto','DescCtto','MonedaCtto','ValorCtto','IdCtta','EstCtto','FechIniCtto','IdCecoCtto','CordCtto','IdMandante' ]
    #Con esta linea establecemos que se hara despues que la operacion de modificacion se complete correctamente
    #success_url = reverse_lazy('personas:personas')

    #def get_success_url(self):
        #Aux2 = Ctto.objects.get(id=self.kwargs['id_ctto']).NumCtto
        #print("Aux3")
        #print (edp.pk)
        #success_url = reverse('personas:EditarContrato',kwargs={'id_ctto': Aux2 })
        #success_url = reverse_lazy('personas:personas')
    def get_success_url(self):
        Aux2 = Edp.objects.get(id=self.kwargs['pk']).IdCtto.NumCtto
        #print( Ctto.objects.get(id=int(self.kwargs['id_ctto']))).NumCtto
        return reverse('personas:EditarContrato',kwargs={'id_ctto': Aux2 })

class CrearEdp(CreateView):
    model = Edp
    #fields =['dni','nombre','apellido_paterno','apellido_materno']
    template_name = 'crear_edp_new.html'
    form_class = EdpCreateForm
    success_url = reverse_lazy('personas:EditarContrato')
    D_edp = Edp


    def get_context_data(self, **kwargs):
        context = super(CrearEdp, self).get_context_data(**kwargs)
        id_ctto = self.kwargs['id_ctto']
        context['Valedp'] = Edp.objects.all()
        context['Validctto'] = int(self.kwargs['id_ctto'])
        context['NumeroCtto'] = Ctto.objects.get(id=self.kwargs['id_ctto']).NumCtto
        context['DescripCtto'] = Ctto.objects.get(id=self.kwargs['id_ctto']).DescCtto
        context['ItemCtto'] = ItemCtto.objects.filter(IdCtto__id=id_ctto).order_by('NumItem')

        print("valor de idctto =")
        print(id_ctto)
        return context

    def get_form_kwargs(self):
        kwargs = super(CrearEdp, self).get_form_kwargs()
        kwargs.update({'idctto':self.kwargs['id_ctto'],'dato_aux':'dato2'})

        return kwargs

    def get_success_url(self):
        Aux2 = Ctto.objects.get(id=self.kwargs['id_ctto']).NumCtto
        #print( Ctto.objects.get(id=int(self.kwargs['id_ctto']))).NumCtto
        return reverse('personas:EditarContrato',kwargs={'id_ctto': Aux2 })



class BorrarEdp(DeleteView):
    #Especificamos que el modelo a utilizar va a ser Persona
    form_class = EdpUpdateForm

    model = Edp
    #Establecemos que la plantilla se llamara modificar persona
    template_name = 'modificar_Edp_new.html'
    #Determinamos los campos con los que se va a trabajar, esto es obligatorio sino nos saldra un error
    #fields = ['NumCtto','DescCtto','MonedaCtto','ValorCtto','IdCtta','EstCtto','FechIniCtto','IdCecoCtto','CordCtto','IdMandante' ]
    #Con esta linea establecemos que se hara despues que la operacion de modificacion se complete correctamente
    success_url = reverse_lazy('personas:personas')

    def get_context_data(self, **kwargs):
        context = super(BorrarEdp, self).get_context_data(**kwargs)

        context['NumeroCtto'] = Edp.objects.get(id=self.kwargs['pk']).IdCtto.NumCtto
        return context

    def get_success_url(self):
        Aux2 = Edp.objects.get(id=self.kwargs['pk']).IdCtto.NumCtto
        #print( Ctto.objects.get(id=int(self.kwargs['id_ctto']))).NumCtto
        return reverse('personas:EditarContrato',kwargs={'id_ctto': Aux2 })




class ModificarOdc(UpdateView):
    #Especificamos que el modelo a utilizar va a ser Persona
    form_class = OdcUpdateForm

    model = Odc
    #Establecemos que la plantilla se llamara modificar persona
    template_name = 'modificar_odc_new.html'
    #Determinamos los campos con los que se va a trabajar, esto es obligatorio sino nos saldra un error
    #fields = ['NumCtto','DescCtto','MonedaCtto','ValorCtto','IdCtta','EstCtto','FechIniCtto','IdCecoCtto','CordCtto','IdMandante' ]
    #Con esta linea establecemos que se hara despues que la operacion de modificacion se complete correctamente
    success_url = reverse_lazy('personas:personas')

    #def get_success_url(self):
        #Aux2 = Ctto.objects.get(id=self.kwargs['id_ctto']).NumCtto
        #print("Aux3")
        #print (edp.pk)
        #success_url = reverse('personas:EditarContrato',kwargs={'id_ctto': Aux2 })
        #success_url = reverse_lazy('personas:personas')
    def get_context_data(self, **kwargs):
        data = super(ModificarOdc, self).get_context_data(**kwargs)
        if self.request.POST:
            data['familymembers'] = ItemOdcFormSet(self.request.POST, instance=self.object)
        else:
            data['familymembers'] = ItemOdcFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        familymembers = context['familymembers']
        with transaction.atomic():
            self.object = form.save()

            if familymembers.is_valid():
                familymembers.instance = self.object
                familymembers.save()
        return super(ModificarOdc, self).form_valid(form)


    def get_success_url(self):
        Aux2 = Odc.objects.get(id=self.kwargs['pk']).IdCtto.NumCtto
        #print( Ctto.objects.get(id=int(self.kwargs['id_ctto']))).NumCtto
        return reverse('personas:EditarContrato',kwargs={'id_ctto': Aux2 })








class CrearOdc(CreateView):
    model = Odc
    #fields =['dni','nombre','apellido_paterno','apellido_materno']
    template_name = 'crear_odc_new.html'
    form_class = OdcCreateForm
    success_url = reverse_lazy('personas:EditarContrato')
    D_edp = Edp


    def get_context_data(self, **kwargs):
        data = super(CrearOdc, self).get_context_data(**kwargs)
        data['Valodc'] = Odc.objects.all()
        data['Validctto'] = int(self.kwargs['id_ctto'])
        data['NumeroCtto'] = Ctto.objects.get(id=self.kwargs['id_ctto']).NumCtto
        data['DescripCtto'] = Ctto.objects.get(id=self.kwargs['id_ctto']).DescCtto

        #return context

        #data = super(CrearOdc, self).get_context_data(**kwargs)
        #print("imprimiendo Data")
        #print(data)

        if self.request.POST:
            data['ItemOdcs'] = ItemOdcFormSet(self.request.POST)
        else:
            data['ItemOdcs'] = ItemOdcFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        ItemOdcs = context['ItemOdcs']
        with transaction.atomic():
            self.object = form.save()

            if ItemOdcs.is_valid():
                ItemOdcs.instance = self.object
                ItemOdcs.save()
        return super(CrearOdc, self).form_valid(form)



    def get_form_kwargs(self):
        kwargs = super(CrearOdc, self).get_form_kwargs()
        kwargs.update({'idctto':self.kwargs['id_ctto'],'dato_aux':'dato2'})

        return kwargs

    def get_success_url(self):
        Aux2 = Ctto.objects.get(id=self.kwargs['id_ctto']).NumCtto
        crear_docODC(kwargs={'id_ctto': Aux2 })
        #print( Ctto.objects.get(id=int(self.kwargs['id_ctto']))).NumCtto
        return reverse('personas:EditarContrato',kwargs={'id_ctto': Aux2 })



class BorrarOdc(DeleteView):
    #Especificamos que el modelo a utilizar va a ser Persona
    form_class = OdcUpdateForm

    model = Odc
    #Establecemos que la plantilla se llamara modificar persona
    template_name = 'modificar_Odc_new.html'
    #Determinamos los campos con los que se va a trabajar, esto es obligatorio sino nos saldra un error
    #fields = ['NumCtto','DescCtto','MonedaCtto','ValorCtto','IdCtta','EstCtto','FechIniCtto','IdCecoCtto','CordCtto','IdMandante' ]
    #Con esta linea establecemos que se hara despues que la operacion de modificacion se complete correctamente
    success_url = reverse_lazy('personas:personas')

    def get_success_url(self):
        Aux2 = Odc.objects.get(id=self.kwargs['pk']).IdCtto.NumCtto
        #print( Ctto.objects.get(id=int(self.kwargs['id_ctto']))).NumCtto
        return reverse('personas:EditarContrato',kwargs={'id_ctto': Aux2 })


class DetalleOdc(DetailView):
    model = Odc
    template_name = 'detalle_persona_new.html'







from django.core import serializers


class BusquedaAjaxView(TemplateView):

    def get(self, request, *arg, **kwargs):
        id_ceco =request.GET['idajx']
        nombre_ceco = Ceco.objects.filter(id=id_ceco)
        data = serializers.serialize('json',nombre_ceco,
                    fields=('NomCeco'))
        return HttpResponse(data, content_type ='application/json')

class CrearContratista(CreateView):
        model = Ctta
        #fields =['dni','nombre','apellido_paterno','apellido_materno']
        template_name = 'crear_ctta_new.html'
        form_class = CttaUpdateForm
        success_url = reverse_lazy('personas:crear_contrato')




class Crear_Personalproy(CreateView):

    model = PersonalProyecto
    #fields =['dni','nombre','apellido_paterno','apellido_materno']
    template_name = 'crear_personalproy_new.html'
    form_class = PersonalProyUpdateForm

    success_url = reverse_lazy('personas:crear_contrato')




class Crear_Personalctta(CreateView):
        model = PersonalCtta
        #fields =['dni','nombre','apellido_paterno','apellido_materno']
        template_name = 'crear_personalctta_new.html'
        form_class = PersonalCttaUpdateForm
        success_url = reverse_lazy('personas:crear_contrato')
